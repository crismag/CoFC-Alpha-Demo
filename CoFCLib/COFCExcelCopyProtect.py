import pandas as pd
import openpyxl
import xlrd
import xlwt

class COFCCopyProtect:
    def __init__(self, excel_infile, excel_outfile, sheets_of_interest=None, use_pandas_only=True):
        self.excel_infile = excel_infile
        self.excel_outfile = excel_outfile
        self.sheets_of_interest = sheets_of_interest
        self.use_pandas_only = use_pandas_only

    def copy_worksheets(self):
        if self.use_pandas_only:
            self.copy_with_pandas()
        else:
            self.copy_with_xlrd_openpyxl()

    def copy_with_pandas(self):
        source_excel = pd.ExcelFile(self.excel_infile)
        sheet_names = source_excel.sheet_names if not self.sheets_of_interest else self.sheets_of_interest

        with pd.ExcelWriter(self.excel_outfile, mode='w') as writer:
            for sheet_name in sheet_names:
                source_df = pd.read_excel(self.excel_infile, sheet_name=sheet_name)
                source_df.to_excel(writer, sheet_name=sheet_name, index=False)

    def copy_with_xlrd_openpyxl(self):
        source_excel_version = self.detect_excel_version(self.excel_infile)

        if source_excel_version == 'xlsx':
            source_workbook = openpyxl.load_workbook(self.excel_infile, data_only=True)
            output_workbook = openpyxl.Workbook()

            for sheet_name in source_workbook.sheetnames:
                if not self.sheets_of_interest or sheet_name in self.sheets_of_interest:
                    source_sheet = source_workbook[sheet_name]
                    output_sheet = output_workbook.create_sheet(title=sheet_name)

                    for row in source_sheet.iter_rows(values_only=True):
                        output_sheet.append(row)

            output_workbook.remove(output_workbook['Sheet'])  # Remove the default sheet
            output_workbook.save(self.excel_outfile)

        elif source_excel_version == 'xls':
            workbook = xlrd.open_workbook(self.excel_infile, formatting_info=True)
            new_workbook = xlwt.Workbook()
            for sheet_index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_index)
                if not self.sheets_of_interest or sheet.name in self.sheets_of_interest:
                    new_sheet = new_workbook.add_sheet(sheet.name)
                    for row in range(sheet.nrows):
                        for col in range(sheet.ncols):
                            new_sheet.write(row, col, sheet.cell(row, col).value)
            new_workbook.save(self.excel_outfile)

    def detect_excel_version(self, file_path):
        if file_path.lower().endswith('.xls'):
            return 'xls'
        elif file_path.lower().endswith('.xlsx'):
            return 'xlsx'
        else:
            raise ValueError("Unsupported Excel file format")

    def get_worksheets(self):
        source_excel = pd.ExcelFile(self.excel_infile)
        return source_excel.sheet_names

# Example usage:
is_test_enable_data_test = 0
if __name__ == "__main__" and is_test_enable_data_test:
    use_pandas_only = False
    excel_infile: str = r'C:\Users\magalang\Documents\COA_M3990B1.protected.xls'
    if use_pandas_only == True:
        excel_outfile = 'unprotected.xlsx'
    else:
        excel_outfile = 'unprotected.xls'
    sheets_of_interest = ['CofC','Statistics']  # Specify worksheets of interest, or None for all
    copier = COFCCopyProtect(excel_infile, excel_outfile, sheets_of_interest, use_pandas_only)
    copier.copy_worksheets()
    worksheets = copier.get_worksheets()
    print("Worksheets:", worksheets)
